# src/analytics/router.py

from datetime import date, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
import io
import openpyxl
from sqlalchemy import func, select
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.analytics.schemas import FailureStats, ForecastResponse
from src.auth.dependencies import get_current_admin_user, get_current_user
from src.part_types.dao import PartTypeDAO
from src.failure_records.dao import FailureRecordDAO
from src.database import async_session_maker
from src.devices.models import Device
from src.device_types.models import DeviceType
from src.maintenance_tasks.models import MaintenanceTask
from src.failure_records.models import FailureRecord
from src.write_off_reports.dao import WriteOffReportDAO

router = APIRouter(
    prefix="/analytics",
    tags=["Аналитика"],
    dependencies=[Depends(get_current_admin_user)],
)


@router.get(
    "/forecast",
    response_model=ForecastResponse,
    summary="Прогноз даты следующей замены по всем деталям",
    description="Возвращает прогноз даты следующей замены детали на основе среднего интервала отказа по всем типам деталей и последнего зафиксированного отказа.",
)
async def forecast_replacement(current_user=Depends(get_current_user)):
    part_types = await PartTypeDAO.find_all()
    intervals = [
        pt.expected_failure_interval_days
        for pt in part_types
        if pt.expected_failure_interval_days
    ]
    if not intervals:
        raise HTTPException(
            status_code=400, detail="Нет данных о средних интервалах отказа"
        )
    avg_interval = int(sum(intervals) / len(intervals))

    records = await FailureRecordDAO.find_all_by_creator_id(creator_id=current_user.id)
    if records:
        last_date = max(r.failure_date for r in records)
    else:
        last_date = date.today()

    forecast = last_date + timedelta(days=avg_interval)
    return ForecastResponse(forecast_replacement_date=forecast)


@router.get(
    "/summary",
    summary="Общая статистика по оборудованию и эксплуатации",
    response_description="JSON с общей статистикой по парку оборудования",
)
async def summary_stats(current_user=Depends(get_current_user)):
    async with async_session_maker() as session:
        # Общее количество устройств
        total_devices = (await session.execute(select(func.count(Device.id)))).scalar()
        # Количество устройств по статусам
        statuses = (
            await session.execute(
                select(Device.status, func.count(Device.id)).group_by(Device.status)
            )
        ).all()
        status_counts = {s: c for s, c in statuses}
        # Количество уникальных типов устройств
        unique_types = (
            await session.execute(select(func.count(DeviceType.id)))
        ).scalar()
        # Количество отказов
        total_failures = (
            await session.execute(select(func.count(FailureRecord.id)))
        ).scalar()
        # Количество обслуживаний по типу
        maints = (
            await session.execute(
                select(
                    MaintenanceTask.task_type, func.count(MaintenanceTask.id)
                ).group_by(MaintenanceTask.task_type)
            )
        ).all()
        maint_counts = {t: c for t, c in maints}
        # Средний возраст устройств (в днях)
        avg_age = (
            await session.execute(
                select(
                    func.avg(
                        func.extract("epoch", func.now() - Device.purchase_date)
                        / 86400.0
                    )
                )
            )
        ).scalar()
        # Количество устройств по производителям
        manufacturers = (
            await session.execute(
                select(DeviceType.manufacturer, func.count(Device.id))
                .join(Device, Device.type_id == DeviceType.id)
                .group_by(DeviceType.manufacturer)
            )
        ).all()
        manufacturer_counts = {m: c for m, c in manufacturers}
        # Количество списанных устройств
        decommissioned_count = (
            await session.execute(
                select(func.count(Device.id)).where(Device.status == "decommissioned")
            )
        ).scalar()
        # Количество отчётов о списании
        writeoff_reports_count = len(await WriteOffReportDAO.find_all())
    return {
        "total_devices": total_devices,
        "status_counts": status_counts,
        "unique_device_types": unique_types,
        "total_failures": total_failures,
        "maintenance_counts": maint_counts,
        "avg_device_age_days": round(avg_age, 1) if avg_age else None,
        "devices_by_manufacturer": manufacturer_counts,
        "decommissioned_devices": decommissioned_count,
        "writeoff_reports_count": writeoff_reports_count,
    }


@router.get(
    "/summary-xlsx",
    summary="Выгрузка общей статистики по оборудованию в xlsx",
    response_description="Excel-файл с общей статистикой по парку оборудования",
)
async def summary_stats_xlsx(current_user=Depends(get_current_user)):
    async with async_session_maker() as session:
        total_devices = (await session.execute(select(func.count(Device.id)))).scalar()
        statuses = (
            await session.execute(
                select(Device.status, func.count(Device.id)).group_by(Device.status)
            )
        ).all()
        status_counts = {s: c for s, c in statuses}
        unique_types = (
            await session.execute(select(func.count(DeviceType.id)))
        ).scalar()
        total_failures = (
            await session.execute(select(func.count(FailureRecord.id)))
        ).scalar()
        maints = (
            await session.execute(
                select(
                    MaintenanceTask.task_type, func.count(MaintenanceTask.id)
                ).group_by(MaintenanceTask.task_type)
            )
        ).all()
        maint_counts = {t: c for t, c in maints}
        avg_age = (
            await session.execute(
                select(
                    func.avg(
                        func.extract("epoch", func.now() - Device.purchase_date)
                        / 86400.0
                    )
                )
            )
        ).scalar()
        manufacturers = (
            await session.execute(
                select(DeviceType.manufacturer, func.count(Device.id))
                .join(Device, Device.type_id == DeviceType.id)
                .group_by(DeviceType.manufacturer)
            )
        ).all()
        manufacturer_counts = {m: c for m, c in manufacturers}
        decommissioned_count = (
            await session.execute(
                select(func.count(Device.id)).where(Device.status == "decommissioned")
            )
        ).scalar()
        writeoff_reports_count = len(await WriteOffReportDAO.find_all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Стили
    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    section_fill = PatternFill("solid", fgColor="BDD7EE")
    center_align = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    # Заголовок
    ws.merge_cells("A1:B1")
    ws["A1"] = "Общая статистика по парку оборудования"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    row = 3

    def add_section(title, data):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = bold_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).alignment = center_align
        row += 1
        for k, v in data:
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1

    # Основные метрики
    add_section(
        "Основные показатели",
        [
            ("Общее количество устройств", total_devices),
            ("Уникальных типов устройств", unique_types),
            ("Общее количество отказов", total_failures),
            (
                "Средний возраст устройств (дней)",
                round(avg_age, 1) if avg_age else None,
            ),
        ],
    )

    # По статусам
    add_section("Устройства по статусам", status_counts.items())

    # По обслуживанию
    add_section("Обслуживания по типу", maint_counts.items())

    # По производителям
    add_section("Устройства по производителям", manufacturer_counts.items())

    # Списанное оборудование
    add_section(
        "Списанное оборудование",
        [
            ("Списанных устройств", decommissioned_count),
            ("Отчётов о списании", writeoff_reports_count),
        ],
    )

    # Границы для всех ячеек с данными
    for r in ws.iter_rows(min_row=3, max_row=row - 1, min_col=1, max_col=2):
        for cell in r:
            cell.border = border

    # Автоширина
    for col_idx, col in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1
    ):
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=summary_stats.xlsx"},
    )
